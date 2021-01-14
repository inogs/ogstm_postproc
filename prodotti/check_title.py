import argparse

def argument():
    parser = argparse.ArgumentParser(description = '''
    Checks consistency between NetCDF product file and PIT about title
   ''',formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument(   '--filename', '-f',
                                type = str,
                                required = True,
                                help ='Path of the NetCDF product file'
                                )
    
    parser.add_argument(   '--dataset',"-d",
                                type = str,
                                required = True,
                                help = 'dataset name')
    parser.add_argument(   '--pit',"-p",
                                type = str,
                                required = True,
                                help = 'Path of the PIT file')
    return parser.parse_args()


args = argument()    


from openpyxl import load_workbook
import netCDF4

PIT_file = args.pit
dataset = args.dataset
filename=args.filename


wb = load_workbook(filename = PIT_file, read_only=True, data_only=True)
work_sheet = wb['Dataset']

for row in range(2,200):
    if work_sheet.cell(row=row, column=2).value is not None:
        xls_dataset = work_sheet.cell(row=row, column=2).value
        if xls_dataset == dataset:
            xls_title = work_sheet.cell(row=row, column=18).value
            break

print xls_title

D=netCDF4.Dataset(filename,"r")
nc_title=str(getattr(D,'title'))
D.close()

if xls_title == nc_title : 
    print "OK"
else:
    print "ERROR: inconsistency between NetCDF and PIT"
